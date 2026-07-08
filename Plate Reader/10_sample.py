import random
import argparse
import pandas as pd
import numpy as np
import openpyxl as opx
from scipy.integrate import simpson
from scipy.integrate import trapezoid
from scipy.integrate import cumulative_trapezoid
from scipy.interpolate import interp1d
from datetime import date
from openpyxl import load_workbook

#rippable end spike calculator
def spk_calc(x):
    if x<=0:
        result= start_dil/2
    elif x <= 1:
        result= start_dil
    else:
        result=((start_dil)*((dil_fact))**(x-1))
    return result

# arg parser for the script
if __name__ == '__main__':
    parser = argparse.ArgumentParser(prog='10 sample elisa plate analysis', 
        description= 'Feed in elisa plate data spit out fully analyzed data',
            epilog="||||| Ver 1.1.0: Martini ||||| Latest changes: Rework AUC calculation and error handling")
    parser.add_argument('filepath', help='Filename of the file to be processed *CASE SENSITIVE* .CSV or XLSX format only!')
    parser.add_argument('-us', '--update_source', help='Update the file that you are accessing the data from', action='store_true')
    parser.add_argument('-rp', '--replicate', help='Replicate number how many replicates of one sample are present in the plate. * Only whole numbers between 1-8 *', action='store', default=1)
    parser.add_argument('-bp', '--blankpos', help='Set positions for the blanks in the well for calculation. OPTIONS: s=split, r=right, l=left', action='store', default='')
    parser.add_argument('-sn','--sheetname', help='If data is in a specific sheet Specify sheet name *CASE SENSITIVE*.   Default="Sheet1" ', action='store' ,default='Sheet1')
    parser.add_argument('-i', '--initials', help='Initials of the User.   Default=UNK', action='store', default="Unk")
    parser.add_argument('-sd', '--start_dil', help="Starting dilution #'s only.   Default = 100", action='store', default=100)
    parser.add_argument('-c', '--cut_off', help="Cutoff for your analysis.  Default=0.15", action='store', default=0.15)
    parser.add_argument('-df', '--dil_fact', help="Dilution factor of assay whole #'s only and no 0's.  Default = 2", action='store', default=2)
    parser.add_argument('-r', '--round', help='Change the number of decimal points. Default = 2', action='store', default=2)
    parser.add_argument('-of', '--outfile', help='Custom destination of output file. Default value only works on mac with sharepoint synced!', 
                    action='store',
                    
                    # Want to modify the Output to work automatically EDIT 1 LINE BELOW:
                    
                    default='~/The Mount Sinai Hospital/Krammer Lab Serology Core - Documents/Python Output/')

                    #Gone too far!

    args=parser.parse_args()

    r2=[]
    c_end_spk=[]
    auc=[]
    plate_list=[]
    plate_list_2=[]
    plate_list_3=[]
    long_label=[]
    blank_list=[]
    cut_list=[]

    update_source = args.update_source
    rep_count= int(args.replicate)
    dil_fact = int(args.dil_fact)
    start_dil = int(args.start_dil)
    sname = str(args.sheetname)
    rnd = int(args.round)
    blanks = str(args.blankpos)
    plate_skip=[]
    
    xvals=[(1/2)*start_dil]

    # Calculation of titer values DERIVED FROM df and sd
    i=0
    while i<12:
        if i == 0:
            xvals.append(xvals[i]*2)
            i+=1
        else:
            xvals.append(xvals[i]* dil_fact)
            i+=1

    print(xvals)

#add more functionality here
#right now takes Excel and CSV
    try:
        plates_from_file = pd.read_excel(args.filepath, sheet_name=args.sheetname, header=None)
    except:
        try:
            plates_from_file = pd.read_csv(args.filepath, header=None, sep=',')
        except:

            print("ERROR TYPE: Origin Filetype is not an acceptable type!")
            print(".csv and .xlsx files supported!")
            quit()

    #drops Unnecessary cells from plate
    pff_ind = plates_from_file.index
    plates_from_file.dropna(axis=0, thresh=4, inplace=True)
    plates_from_file.dropna(axis=1, thresh=0.6*len(plates_from_file.index), inplace=True)
    try:
        plates_from_file.drop(pff_ind[-1], axis=1, inplace=True)
    except:
        try:
            plates_from_file.drop(pff_ind[14],axis=1,inplace=True)
        except:
            pass

    plates_from_file.index = np.arange(plates_from_file.shape[0])
    
    #extracting the specific plates from the sheet
    plate_list_raw = [plates_from_file.iloc[9*i:9*(i+1), : ].copy().rename(index=lambda val: val % 9) for i in range(plates_from_file.shape[0] // 9)]

    for plate in plate_list_raw:
        
        plate.drop(plate.columns[0], axis=1, inplace=True)
        platemod = plate.drop(0, axis=0)
        if blanks != '':
            if blanks == 's' or blanks == 'S':
                blank_agg_1 = platemod.iloc[:,0].to_list()
                blank_agg_2 = platemod.iloc[:,-1].to_list()
                blank_agg = blank_agg_1 + blank_agg_2
            elif blanks == 'r' or blanks == 'R':
                blank_agg_1 = platemod.iloc[:,-1].to_list()
                blank_agg_2 = platemod.iloc[:,-2].to_list()
                blank_agg = blank_agg_1 + blank_agg_2
            elif blanks == 'l' or blanks == 'L':
                blank_agg_1 = platemod.iloc[:,0].to_list()
                blank_agg_2 = platemod.iloc[:,1].to_list()
                blank_agg = blank_agg_1 + blank_agg_2
            blank_list.append(blank_agg)


    #tranposition to horizontal for 10_sample plate 
    for plate in plate_list_raw:
        plate = plate.T
        plate.index=np.arange(plate.shape[0])
        plate_list.append(plate)

    #Plate processing and label catching
    n1=0
    for plate in plate_list:
        
        overflow = 0

        ind = plate.index
        col = plate.columns

        index_label = plate.iloc[:,0]
        long_label = long_label + index_label.to_list()

        plate.drop(col[0], axis=1, inplace=True)
        
        if blanks != '':
            if blanks == 's' or blanks == 'S':
                plate.drop(ind[0], axis=0, inplace=True)
                plate.drop(ind[-1], axis=0, inplace=True)
              
            elif blanks == 'r' or blanks == 'R':
                plate.drop(ind[0], axis=0, inplace=True)
                plate.drop(ind[1], axis=0, inplace=True)
               
            elif blanks == 'l' or blanks == 'L':
                plate.drop(ind[-1], axis=0, inplace=True)
                plate.drop(ind[-2], axis=0, inplace=True)
             

        ovrflw = plate[plate == 'OVRFLW'].count()
        overflow = ovrflw[ovrflw > 0].count()

        if overflow >= 1:
            print('Over Flow Error! Skipping Plate '+ str(n1+1))
            plate_skip.append(n1)
            n1+=1
            continue
        
        n1+=1

        print(plate.iloc[:,1])

        a1 = sum(plate.iloc[:,1])/len(plate.iloc[:,1])
        a2 = sum(plate.iloc[:,2])/len(plate.iloc[:,2])
        if a1 < a2:
            plate.drop(plate.columns[0], axis=1, inplace=True)

        plate_list_2.append(plate)

    for item in blank_list:
        avg = sum(item)/len(item)
        var = sum([((x-avg) ** 2) for x in item])/ (len(item) - 1)
        res = var**0.5
        cut_list.append(np.round((avg+(5*res)),9))

    if len(plate_skip) == len(plate_list):
        print("All plates invalid! Review source file!")
        print("Program Stopped Early!")
        quit()

    n2=0

    for plate in plate_list_2:
        
        column = plate.columns
        
        if blanks != '':
            cut = cut_list[n2]
            n2+=1
        else:
            cut=float(args.cut_off)

        if rep_count != 0:
            if blanks != 'l':
                plate_2 = plate.groupby((plate.index - 1) // rep_count).mean().round(4)
            else:
                plate_2 = plate.groupby((plate.index) // rep_count).mean().round(4)
        else:
            plate_2 = plate

        column = plate_2.columns

        mapper = {col: i for i, col in enumerate(plate.columns)}
        # print("index", plate.index)
        # print("columns", plate.columns)
        # plate3 = plate2.drop(0,axis=1)
        plate_2['Over_Cut'] = (plate_2.iloc[:,:-1] >= cut).sum(axis=1)
        plate_2['Dilutions'] = ((plate_2.iloc[:, :-2] <= cut).idxmax(axis=1).apply(lambda col: mapper[col]))
        plate_2['End_Titer'] = plate_2['Dilutions'].apply(spk_calc)
        plate_2['Pre Flag'] = ((plate_2['Over_Cut']) - (plate_2['Dilutions']))
        plate_2.loc[plate_2['Pre Flag'] < 0.0, "Flag"] = "SYS ERR"
        plate_2.loc[plate_2['Pre Flag'] <= 7.0, "Flag"] = "Check Trend"
        plate_2.loc[plate_2['Pre Flag'] == 0.0, "Flag"] = "Ok"
        plate_2.loc[plate_2['Pre Flag'] >= 10, "Flag"] = "Titer High"
        plate_2.drop('Pre Flag', inplace=True, axis=1)
        plate_list_3.append(plate_2)

    # REALLY need to reevaluate this at a later date
    # AUC and Continuous End spike calculation

    n3=0    

    for plate in plate_list_3:
        
        if blanks != '':
            cut = cut_list[n3]
            n3 += 1
        else:
            cut=float(args.cut_off)

        auc=[]
        c_end_spk=[]
        table_cut=[]
        for name,row in plate.iterrows():
            print('')
            table_cut.append(cut)
            i=0
            r2=[]
            if row.iloc[-4] > cut:
                if row.iloc[-5] > cut:
                    auc.append('N/A')
                    c_end_spk.append('N/A')
                    continue

            r = row[:-4].tolist()

            for val in r:
                r2.append(val-cut)
            
            if row['Dilutions'] == 0 or row['Dilutions'] != row['Dilutions']:
                auc.append('N/A')
                c_end_spk.append('N/A')
                continue

            intpx = [xvals[int(row['Dilutions'])], xvals[int(row['Dilutions']+1)]]
            intpy = [r2[int(row['Dilutions']-1)], r2[int(row['Dilutions'])]]

            if intpy[0] > 0 and intpy[1] < 0:
                slope = (intpy[0]-intpy[1])/(intpx[0]-intpx[1])
                print('Slope: ', slope)
                b = intpy[0]-(intpx[0]*slope)
                print('Y intercept: ', b)
                ycross = (0-b)/slope
                print('0 crossing: ', ycross)
                
                c_end_spk.append(ycross.round(4))
                trend = r2[:int(row['Dilutions'])]
                newxvals = xvals[1:len(trend)+1]
                newxvals.append(float(ycross))
                trend.append(0)
                auc.append(round(trapezoid(trend,newxvals), rnd))

            else:
                print('Interpolation Error!: no 0 crossing within test points')
                print('X values: ', intpx)
                print('Y values: ',intpy)
                print('Full Trend:')
                print(row)

                auc.append('Inter Error')
        
        plate['AUC'] = auc
        plate['Inter_Titer'] = c_end_spk
        plate['Cut Off'] = table_cut
        
    AUC_final=pd.concat(plate_list_3, axis = 0)

    label_beg = xvals[1:len(AUC_final.columns)-6]
    label_end = AUC_final.columns[-7:].tolist()
    label_list = label_beg + label_end
    AUC_final.columns = label_list    
    
    td = date.today()
    td=td.strftime("%d_%m_%Y")
    if update_source == True:
        try:
            book = load_workbook(args.filepath)
            writer = pd.ExcelWriter(args.filepath, engine= 'openpyxl')
            writer.book = book
            AUC_final.to_excel( writer , sheet_name='Analyzed Data')
            writer.close()
        except:
             AUC_final.to_csv(args.outfile + '8_sample_plate_' + args.initials + '_' + td + '.csv', sep=',')
             print('Filepath not Extended!')
             print('Please Extend The Filepath to Base Filepath')
             print('MAC= Users/')
             print('PC= C:/')
             quit()
    else:
        try:
             AUC_final.to_csv(args.outfile , sep=',')
        except:
            AUC_final.to_csv(args.outfile + '8_sample_plate_' + args.initials + '_' + td + '.csv', sep=',')


    out_mess = random.randint(0,100)

    if out_mess <= 25:
        print("What's the meaning of life?")
    elif out_mess <= 50:
        print("Done already?, Yea I'm that good ;)")
    elif out_mess <= 60:
        print("I'm finished, you can stop staring")
    elif out_mess <= 70:
        print("#daayyuuuummm")
    elif out_mess <= 80:
        print("Jack of no trades master of one!")
    elif out_mess <= 85:
        print("DiD iT fInIsH yEt?!?!?")
    elif out_mess <= 90:
        print("When your brain doesn't work like it used to before")
    elif out_mess <= 95:
        print("Would you like cheese with that wine?")
    elif out_mess <= 99:
        print("SysError: Not Enough tomato")
    elif out_mess <= 100:
        print("42! The Best Answer!")

    # try:
    #     book = opx.load_workbook(args.filepath)
    #     writer = pd.ExcelWriter(args.filepath, engine='openpyxl')
    #     writer.book = book
    #     AUC_final.to_excel(writer, sheet_name='Py Process')
    #     writer.close()
    # except:
    #     try:
    #         AUC_final.to_csv(args.outfile, sep=',')
    #     except:
    #         AUC_final.to_csv(args.outfile + '10_sample_plate_' + args.initials + '_' + td + '.csv', sep=',')

