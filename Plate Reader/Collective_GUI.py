import sys
import PySimpleGUI as sg
import argparse
import random
import pandas as pd
import numpy as np
import openpyxl as opx
from openpyxl import load_workbook
from scipy.integrate import trapezoid
from scipy.interpolate import interp1d
from datetime import date

# PROGRESS FROM 3/28/23:
# COMPLETELY FUNCTIONAL YOU GOD! ALL GOOD FOR RELEASE

def spk_calc(x, start, fact):
    if x<=0:
        result= start/2
    elif x <= 1:
        result= start
    else:
        result=((start)*((fact))**(x-1))
    return result

class ValuesToClass(object):
        def __init__(self, values):
            for key in values:
                setattr(self, key, values[key])

if __name__ == '__main__':
    if len(sys.argv) != 1:
        parser = argparse.ArgumentParser(prog='8 sample elisa plate analysis',
            description= 'Feed in elisa plate data spit out AUC and End Titer data',
                epilog="||||| ver 1.1.0: Sazerac ||||| Latest Changes: Rework AUC calculation and error handling")
        parser.add_argument('-ft','--filetype', help='filetype: either 8, 10, or 16 sample plates accepted', required=True)
        parser.add_argument('filepath', help='Filename of the file to be processed *CASE SENSITIVE* .CSV or XLSX format only!', required=True)
        parser.add_argument('-us', '--update_source', help='Update the file that you are accessing the data from', action='store_true')
        parser.add_argument('-rp', '--replicate', help='Replicate number how many replicates of one sample are present in the plate. * Only whole numbers between 1-8 *', action='store', default=1)
        parser.add_argument('-bp', '--blankpos', help='Set positions for the blanks in the well for calculation. OPTIONS: s=split, r=right, l=left', action='store', default='')
        parser.add_argument('-sn','--sheetname', help='*EXCEL ONLY* If data is in a specific sheet specify sheet name *CASE SENSITIVE*', action='store' ,default='Sheet1')
        parser.add_argument('-i', '--initials', help='Initials of the User: *FOR AUTOMATED OUTPUT ONLY*', action='store', default="Unk")
        parser.add_argument('-sd', '--start_dil', help="Starting dilution * #'s only and no 0*", action='store', default=100)
        parser.add_argument('-df', '--dil_fact', help="Dilution factor of assay whole * #'s only and no 0*", action='store', default=2)
        parser.add_argument('-r', '--round', help='Change the number of decimal points. Default = 2', action='store', default=2)
        parser.add_argument('-c', '--cut_off', help="Cutoff for your analysis", action='store', default=0.15)
        parser.add_argument('-xlsx', '--Excel', help='output excel file output', action='store_true')
        parser.add_argument('-csv', '--CSV', help='Comma Delimited file output', action='store_true', Default=True)
        parser.add_argument('-tab', '--Tab',help='Comma Delimited file output', action='store_true')
        parser.add_argument('-op', '--outpath', help='Custom destination of output file. *Default value only works on mac with sharepoint synced!*', 
                        
                        action='store',

                        #If you want the automated save to work change the file path here:
                        default='~/The Mount Sinai Hospital/Krammer Lab Serology Core - Documents/Python Output/')
        
        parser.add_argument('-on', '--outname', help='Custom name for the file that you output', action='store', default='')

        args=parser.parse_args()
        outfile = args.outfile

    else:
        sg.theme('Dark Blue 17')

        layout = [[sg.Text('Enter 2 files to compare')],
                [sg.Radio('8 sample plate', 'RADIO1',key='Eight_Sample',default=True), sg.Radio('10 Sample plate', 'RADIO1', key='Ten_Sample'), sg.Radio('16 Sample plate', 'RADIO1', key='Sixteen_Sample')],
                [sg.Text('File\nName', size=(9, 2)), sg.Input(key='filepath'), sg.FileBrowse()],
                [sg.Text('Sheet\nName', size=(9, 2)), sg.Input(key='sheetname')],
                
                [sg.Text('Starting\nDilution', size=(9,2)), sg.Input(size=(4,1), key='start_dil'),\
                  sg.Text('Dilution\nFactor', size=(6,2)), sg.Input(size=(3,1),key='dil_fact'),\
                   sg.Text('Replicates',size=(8,1)), sg.Input(size=(3,1),key='replicate'),\
                    sg.Text('Cut Off'), sg.Input(size=(4,1),key='cut_off')],
                
                [sg.Text('Blank Position'), sg.Radio('2 Left Columns', 'RADIO2', key='Left', default=False),\
                    sg.Radio('Split Left and Right', 'RADIO2', key='Split', default=False),\
                        sg.Radio('2 Right Columns', 'RADIO2', key='Right', default=False)],
                
                [sg.Radio('Manual Cutoff', 'RADIO3', key='MC', default=True),sg.Radio('Blank Calculated', 'RADIO3', key='BC', default=False)],

                [sg.Radio('Update Source File', 'RADIO4', key='update_source'), sg.Radio('Output New File', 'RADIO4', default=True, key="new_file")],
                [sg.Text('Output\nDestination', size=(9, 2)), sg.Input(key='outpath'), sg.FolderBrowse()],
                [sg.Text('Output\nFile Name', size=(9,2)), sg.Input(key='outname')],
                [sg.Radio('Excel file', 'RADIO5',key='Excel'), sg.Radio('Comma Delimited', 'RADIO5', key='CSV', default=True), sg.Radio('Tab Delimited', 'RADIO5', key='Tab')],
                [sg.Submit(), sg.Cancel()]]

        window = sg.Window('File Compare', layout)

        event, values = window.read()
        window.close()

        print(values)

        if event == 'Cancel':
            quit()
        else:
            args = ValuesToClass(values)
            
    try:
        Path_n_Name = args.outpath + '/' + args.outname
        if args.Excel == True:
            outfile = Path_n_Name + '.xlsx'
        elif args.Tab == True:
            outfile = Path_n_Name + '.txt'
        elif args.CSV == True:
            outfile = Path_n_Name + '.csv'
    except:
        Path_n_Name = args.outpath + '/' + args.outname
        if args.Excel == True:
            outfile = Path_n_Name + '.xlsx'
        elif args.Tab == True:
            outfile = Path_n_Name + '.txt'
        elif args.CSV == True:
            outfile = Path_n_Name + '.csv'
    
    r2=[]
    c_end_spk=[]
    auc=[]
    plate_list=[]
    plate_list_t=[]
    plate_list_2=[]
    plate_list_3=[]
    long_label=[]
    blank_list=[]
    cut_list=[]


    if len(sys.argv) != 1:
        if str(args.filetype) == '8s':
            Eight_Sample = True
        elif str(args.filetype) == '10s':
            Ten_Sample = True
        elif str(args.filetype) == '16s':
            Sixteen_Sample = True
        else:
            print("Enter a valid file type:")
            print('8s = 8 Sample', '10s = 10 Sample', '16s = 16 Sample')
            quit()
    else:
        Eight_Sample = args.Eight_Sample
        Ten_Sample = args.Ten_Sample
        Sixteen_Sample = args.Sixteen_Sample

    if len(sys.argv) != 1:
        blanks = str(args.blankpos)
    else:
        if args.Left == True:
            blanks = 'L'
        elif args.Right == True:
            blanks = 'R'
        elif args.Split == True:
            blanks = 'S'
        else:
            blanks = ''
    
    update_source = args.update_source
    
    rep_count= int(args.replicate)
    
    dil_fact = int(args.dil_fact)
    
    start_dil = int(args.start_dil)
    
    sname = str(args.sheetname)
    
    Blank_Calc = args.BC

    try:
        rnd = int(args.round)
    except:
        print('Rounding Set to Default')
        rnd=2
        
    plate_skip=[]
    
    xvals=[(1/2)*start_dil]
    
    i=0
    while i<12:
        if i == 0:
            xvals.append(xvals[i]*2)
            i+=1
        else:
            xvals.append(xvals[i]* dil_fact)
            i+=1

    print(xvals) 
#add more functionality here

    try:
        plates_from_file = pd.read_excel(args.filepath, sheet_name=sname, header=None)
    except:
        try:
            plates_from_file = pd.read_csv(args.filepath, header=None, sep=',')
        except:

            print("ERROR TYPE: Origin Filetype is not an acceptable type!")
            print(".csv and .xlsx files supported!")
            quit()
    

    if Eight_Sample == True:
        plates_from_file.dropna(axis=0, thresh=4, inplace=True)
        plates_from_file.dropna(axis=1, thresh=0.6*len(plates_from_file.index), inplace=True)
        
        # try:
        #     plates_from_file.drop(pff_ind[-1], axis=1, inplace=True)
        # except:
        #     try:
        #         plates_from_file.drop(pff_ind[14],axis=1,inplace=True)
        #     except:
        #         pass
        
        plates_from_file.index = np.arange(plates_from_file.shape[0])
        
        plate_list = [plates_from_file.iloc[9*i:9*(i+1), : ].copy().rename(index=lambda val: val % 9) for i in range(plates_from_file.shape[0] // 9)]
        
        n1=0
        for plate in plate_list:
            print(plate)    
            overflow = 0

            ind = plate.index
            col = plate.columns

            plate.drop(ind[0], axis=0, inplace=True)
            
            if len(plate.columns) == 13:
                plate.drop(col[0], axis=1,inplace=True)

            index_label = plate.iloc[:,0]
            long_label = long_label + index_label.to_list()

            ovrflw = plate[plate == 'OVRFLW'].count()
            overflow = ovrflw[ovrflw > 0].count()

            if overflow >= 1:
                print('Over Flow Error! Skipping Plate '+ str(n1+1))
                plate_skip.append(n1)
                n1+=1
                continue

            n1+=1

            if blanks != '':
                if blanks == 's' or blanks == 'S':
                    blank_agg_1 = plate.iloc[:,0].to_list()
                    blank_agg_2 = plate.iloc[:,-1].to_list()
                    blank_agg = blank_agg_1 + blank_agg_2
                elif blanks == 'r' or blanks == 'R':
                    blank_agg_1 = plate.iloc[:,-1].to_list()
                    blank_agg_2 = plate.iloc[:,-2].to_list()
                    blank_agg = blank_agg_1 + blank_agg_2
                elif blanks == 'l' or blanks == 'L':
                    blank_agg_1 = plate.iloc[:,0].to_list()
                    blank_agg_2 = plate.iloc[:,1].to_list()
                    blank_agg = blank_agg_1 + blank_agg_2
                blank_list.append(blank_agg)
            else:
                a1 = sum(plate.iloc[:,1])/len(plate.iloc[:,1])
                a2 = sum(plate.iloc[:,2])/len(plate.iloc[:,2])
                if a1 < a2:
                    plate.drop(plate.columns[0], axis=1, inplace=True)

            plate_list_2.append(plate)

        if Blank_Calc == True:
            for item in blank_list:
                avg = sum(item)/len(item)
                var = sum([((x-avg) ** 2) for x in item])/ (len(item) - 1)
                res = var**0.5
                cut_list.append(np.round((avg+(5*res)),9))

        if len(plate_skip) == len(plate_list):
            print("All plates invalid! Review source file!")
            print("Program Stopped Early!")
            quit()

        n2=0
        
        for plate in plate_list_2:

            column = plate.columns
            
            if blanks == 's' or blanks =='S':
                    plate.drop(columns=column[0], inplace=True)
                    plate.drop(columns=column[-1], inplace=True)
            if blanks == 'r' or blanks == 'R':
                    plate.drop(columns=column[-1], inplace=True)
                    plate.drop(columns=column[-2], inplace=True)
            if blanks == 'l' or blanks =='L':
                    plate.drop(columns=column[0], inplace=True)
                    plate.drop(columns=column[1], inplace=True)
            
            if Blank_Calc == True:
                cut = cut_list[n2]
                n2+=1
            else:
                cut=float(args.cut_off)

            if rep_count != 0:
                plate_2 = plate.groupby((plate.index - 1) // rep_count).mean().round(4)
            else:
                plate_2 = plate

            column = plate_2.columns

            mapper = {col: i for i, col in enumerate(plate.columns)}
            # print("index", plate.index)
            # print("columns", plate.columns)
            # plate3 = plate2.drop(0,axis=1)
            plate_2['Over_Cut'] = (plate_2.iloc[:,:] >= cut).sum(axis=1)
            plate_2['Dilutions'] = ((plate_2.iloc[:, :-1] <= cut).idxmax(axis=1).apply(lambda col: mapper[col]))
            plate_2['End_Titer'] = plate_2['Dilutions'].apply(lambda x: spk_calc(x, start_dil, dil_fact))
            plate_2['Pre Flag'] = ((plate_2['Over_Cut']) - (plate_2['Dilutions']))
            plate_2.loc[plate_2['Pre Flag'] < 0.0, "Flag"] = "SYS ERR"
            plate_2.loc[plate_2['Pre Flag'] <= 7.0, "Flag"] = "Check Trend"
            plate_2.loc[plate_2['Pre Flag'] == 0.0, "Flag"] = "Ok"
            plate_2.loc[plate_2['Pre Flag'] >= 10, "Flag"] = "Titer High"
            plate_2.drop('Pre Flag', inplace=True, axis=1)
            plate_list_3.append(plate_2)

        

    if Ten_Sample == True:
        plates_from_file.dropna(axis=0, thresh=4, inplace=True)
        plates_from_file.dropna(axis=1, thresh=0.6*len(plates_from_file.index), inplace=True)
        
        # try:
        #     plates_from_file.drop(pff_ind[-1], axis=1, inplace=True)
        # except:
        #     try:
        #         plates_from_file.drop(pff_ind[14],axis=1,inplace=True)
        #     except:
        #         pass
        
        plates_from_file.index = np.arange(plates_from_file.shape[0])
        
        plate_list = [plates_from_file.iloc[9*i:9*(i+1), : ].copy().rename(index=lambda val: val % 9) for i in range(plates_from_file.shape[0] // 9)]
        
        n1=0
        for plate in plate_list:
            
            overflow = 0

            ind = plate.index
            col = plate.columns

            plate.drop(ind[0], axis=0, inplace=True)
            plate.drop(col[0],axis=1,inplace=True)

            if len(plate.columns) == 13:
                plate.drop(col[-1], axis=1, inplace=True)
            
            index_label = plate.iloc[:,0]
            long_label = long_label + index_label.to_list()

            ovrflw = plate[plate == 'OVRFLW'].count()
            overflow = ovrflw[ovrflw > 0].count()

            if overflow >= 1:
                print('Over Flow Error! Skipping Plate '+ str(n1+1))
                plate_skip.append(n1)
                n1+=1
                continue
            
            n1+=1

            if Blank_Calc == True:
                if blanks == 's' or blanks == 'S':
                    blank_agg_1 = plate.iloc[:,0].to_list()
                    blank_agg_2 = plate.iloc[:,-1].to_list()
                    blank_agg = blank_agg_1 + blank_agg_2
                elif blanks == 'r' or blanks == 'R':
                    blank_agg_1 = plate.iloc[:,-1].to_list()
                    blank_agg_2 = plate.iloc[:,-2].to_list()
                    blank_agg = blank_agg_1 + blank_agg_2
                elif blanks == 'l' or blanks == 'L':
                    blank_agg_1 = plate.iloc[:,0].to_list()
                    blank_agg_2 = plate.iloc[:,1].to_list()
                    blank_agg = blank_agg_1 + blank_agg_2
                blank_list.append(blank_agg)
            else:
                a1 = sum(plate.iloc[:,1])/len(plate.iloc[:,1])
                a2 = sum(plate.iloc[:,2])/len(plate.iloc[:,2])
                if a1 < a2:
                    plate.drop(plate.columns[0], axis=1, inplace=True)

            plate_list_2.append(plate)

        if Blank_Calc == True:
            for plate in plate_list_2:
                plate = plate.T
                plate.index=np.arange(plate.shape[0])
                plate.drop(plate.index[0],axis=0,inplace=True)
                plate.drop(plate.index[-1],axis=0,inplace=True)
                plate_list_t.append(plate)


        for item in blank_list:
            print(item)
            avg = sum(item)/len(item)
            var = sum([((x-avg) ** 2) for x in item])/ (len(item) - 1)
            res = var**0.5
            cut_list.append(np.round((avg+(5*res)),9))

        if len(plate_skip) == len(plate_list):
            print("All plates invalid! Review source file!")
            print("Program Stopped Early!")
            quit()

        n2=0
        for plate in plate_list_t:
            
            column = plate.columns
            
            if Blank_Calc == True:
                cut = cut_list[n2]
                n2+=1
            else:
                cut=float(args.cut_off)


            if rep_count != 0:
                plate_2 = plate.groupby((plate.index - 1) // rep_count).mean().round(4)
            else:
                plate_2 = plate

            column = plate_2.columns

            mapper = {col: i for i, col in enumerate(plate.columns)}
            # print("index", plate.index)
            # print("columns", plate.columns)
            # plate3 = plate2.drop(0,axis=1)
            plate_2['Over_Cut'] = (plate_2.iloc[:,:] >= cut).sum(axis=1)
            plate_2['Dilutions'] = ((plate_2.iloc[:, :-1] <= cut).idxmax(axis=1).apply(lambda col: mapper[col]))
            plate_2['End_Titer'] = plate_2['Dilutions'].apply(lambda x: spk_calc(x, start_dil, dil_fact))
            plate_2['Pre Flag'] = ((plate_2['Over_Cut']) - (plate_2['Dilutions']))
            plate_2.loc[plate_2['Pre Flag'] < 0.0, "Flag"] = "SYS ERR"
            plate_2.loc[plate_2['Pre Flag'] <= 7.0, "Flag"] = "Check Trend"
            plate_2.loc[plate_2['Pre Flag'] == 0.0, "Flag"] = "Ok"
            plate_2.loc[plate_2['Pre Flag'] >= 10, "Flag"] = "Titer High"
            plate_2.drop('Pre Flag', inplace=True, axis=1)
            plate_list_3.append(plate_2)


    if Sixteen_Sample == True:

        plates_from_file.dropna(axis=0, thresh=4, inplace=True)
        plates_from_file.dropna(axis=1, inplace=True)
        plates_from_file.index = np.arange(plates_from_file.shape[0])
        
        plate_list = [plates_from_file.iloc[9*i:9*(i+1), : ].copy().rename(index=lambda val: val % 9) for i in range(plates_from_file.shape[0] // 9)]

        for plate in plate_list:
            plate_1 = plate.iloc[:,2:7]
            plate_2 = plate.iloc[:,7:12]
            plate_list_2.append(plate_1)
            plate_list_2.append(plate_2)

        n=0
        for plate in plate_list_2:
            overflow = 0

            plate.drop(0, axis=0, inplace=True)
            
            ovrflw = plate[plate == 'OVRFLW'].count()
            overflow = ovrflw[ovrflw > 0].count()

            if overflow >= 1:
                print('Over Flow Error! Skipping Plate '+ str(n+1))
                plate_skip.append(n)
                n+=1
                continue

            n+=1

            plate.columns = range(plate.columns.size)
            mapper = {col: i for i, col in enumerate(plate.columns)}
            # print("index", plate.index)
            # print("columns", plate.columns)
            #plate3 = plate2.drop(0,axis=1)
            plate['Over_Cut'] = (plate.iloc[:9,:]>=cut).sum(axis=1)
            plate['Dilutions'] = ((plate.iloc[:9,:-1]<=cut).idxmax(axis=1).apply(lambda col: mapper[col]))
            plate['End_Titer'] = plate['Dilutions'].apply(lambda x: spk_calc(x, start_dil, dil_fact))
            plate['Pre Flag'] = ((plate['Over_Cut']) - plate['Dilutions'])
            plate.loc[plate['Pre Flag'] < 0, "Flag"] = "SYS ERR"
            plate.loc[plate['Pre Flag'] <= 4, "Flag"] = "Check Trend"
            plate.loc[plate['Pre Flag'] == 0, "Flag"] = "Ok"
            plate.loc[plate['Pre Flag'] <= 5, "Flag"] = "Titer High"
            plate.drop('Pre Flag', inplace=True, axis=1)

            plate_list_3.append(plate)

        if len(plate_skip) == len(plate_list):
            print("All plates invalid! Review source file!")
            print("Program Stopped Early!")
            quit()

    
    n3=0 
    
    for plate in plate_list_3:

        if Blank_Calc == True:
            cut = cut_list[n3]
            n3 += 1
        else:
            cut=float(args.cut_off)

        auc=[]
        c_end_spk=[]
        table_cut=[]
        for name,row in plate.iterrows():
            print('')
            table_cut.append(cut)
            i=0
            r2=[]
            if row.iloc[-4] > cut:
                if row.iloc[-5] > cut:
                    auc.append('N/A')
                    c_end_spk.append('N/A')
                    continue

            r = row[:-4].tolist()

            for val in r:
                r2.append(val-cut)
            
            if row['Dilutions'] == 0 or row['Dilutions'] != row['Dilutions']:
                auc.append('N/A')
                c_end_spk.append('N/A')
                continue

            print(xvals)
            intpx = [xvals[int(row['Dilutions'])], xvals[int(row['Dilutions']+1)]]
            print(r2)
            intpy = [r2[int(row['Dilutions']-1)], r2[int(row['Dilutions'])]]

            if intpy[0] > 0 and intpy[1] < 0:
                slope = (intpy[0]-intpy[1])/(intpx[0]-intpx[1])
                print('Slope: ', slope)
                b = intpy[0]-(intpx[0]*slope)
                print('Y intercept: ', b)
                ycross = (0-b)/slope
                print('0 crossing: ', ycross)
                
                c_end_spk.append(round(ycross,4))
                trend = r2[:int(row['Dilutions'])]
                newxvals = xvals[1:len(trend)+1]
                newxvals.append(float(ycross))
                trend.append(0)
                auc.append(round(trapezoid(trend,newxvals), rnd))

            else:
                print('Interpolation Error!: no 0 crossing within test points')
                print('X values: ', intpx)
                print('Y values: ',intpy)
                print('Full Trend:')
                print(row)

                auc.append('Inter Error')
                c_end_spk.append('Inter Error')
            # intlf = interp1d(intpy,intpx)

            # ycross = intlf(0)

        plate['AUC'] = auc
        plate['Inter_Titer'] = c_end_spk
        plate['Cut Off'] = table_cut
        
    AUC_final=pd.concat(plate_list_3, axis = 0)

    label_beg = xvals[1:len(AUC_final.columns)-6]
    label_end = AUC_final.columns[-7:].tolist()
    label_list = label_beg + label_end
    try:
        AUC_final.columns = label_list
    except:
        AUC_final.columns = label_list[0,-2]
    td = date.today()
    td=td.strftime("%d_%m_%Y")

    
    if update_source == True:
        try:
            book = load_workbook(args.filepath)
            writer = pd.ExcelWriter(args.filepath, engine= 'openpyxl')
            writer.book = book
            AUC_final.to_excel( writer , sheet_name='Analyzed Data')
            writer.close()
        except:
            print('Filepath not Extended!')
            print('Please Extend The Filepath to Base Filepath')
            print('MAC= Users/')
            print('PC= C:/')
    
    if args.Excel == True:
        book = opx.Workbook()
        writer = pd.ExcelWriter(outfile, engine= 'openpyxl')
        writer.book = book
        AUC_final.to_excel( writer , sheet_name='Analyzed Data')
        writer.close()
        print('exported to: ', outfile)

    if args.Tab == True:
        AUC_final.to_csv(outfile, sep=' ')
        print('Exported to: ', outfile)

    if args.CSV == True:
        AUC_final.to_csv(outfile , sep=',')
        print('exported to: ', outfile)


    out_mess = random.randint(0,100)

    if out_mess <= 25:
        print("What's the meaning of life?")
    elif out_mess <= 50:
        print("Done already?, Yea I'm that good ;)")
    elif out_mess <= 60:
        print("I'm finished, you can stop staring")
    elif out_mess <= 70:
        print("#daayyuuuummm")
    elif out_mess <= 80:
        print("Jack of no trades master of one!")
    elif out_mess <= 85:
        print("DiD iT fInIsH yEt?!?!?")
    elif out_mess <= 90:
        print("When your brain doesn't work like it used to before")
    elif out_mess <= 95:
        print("would you like cheese with that wine?")
    elif out_mess <= 99:
        print("SysError: Not Enough tomato")
    elif out_mess <= 100:
        print("42! The Best Answer!")
    
    quit()
